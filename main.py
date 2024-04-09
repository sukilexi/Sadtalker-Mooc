import cv2
import os
import openpyxl
import re
import argparse
from moviepy.editor import VideoFileClip, AudioFileClip

#计算每行的字数总和
def calculate_sum(text):
    chinese_count = len(re.findall(r'[\u4e00-\u9fff]', text))
    english_count = len(re.findall(r'[a-zA-Z]', text))
    digit_count = len(re.findall(r'\d', text))
    return chinese_count + english_count + digit_count

# 构建命令行参数解析器
parser = argparse.ArgumentParser(description='Create a video with fade effect using images')
parser.add_argument('--video_file', type=str, help='Path to the video file')
parser.add_argument('--image_file', type=str, help='Path to the image folder')
parser.add_argument('--excel_file', type=str, help='Path to the excel file')
parser.add_argument('--save_name', type=str, help='savename')
args = parser.parse_args()


# 视频文件名和图片文件夹路径
video_file = args.video_file
image_folder = args.image_file
excel_file = args.excel_file
save_name = args.save_name

# 读取Excel表格
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

# 获取表格的行数
num_rows = sheet.max_row

#存储结果
results = []
for row in range(2, num_rows + 1):
    cell_value = sheet.cell(row=row, column=2).value
    if cell_value:
        total_sum = calculate_sum(cell_value)
        results.append(total_sum)
word_sum = 0
for result in results:
    word_sum = word_sum + result

# 读取视频
cap = cv2.VideoCapture(video_file)
fps = cap.get(cv2.CAP_PROP_FPS)
total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
duration = total_frames / fps

# 读取图片文件夹中的图片
images = [img for img in os.listdir(image_folder) if img.endswith(".png")]
sorted_images = sorted(images, key=lambda x: int(x.split('.')[0]))
num_images = len(images)

# 计算每个图片应该持续的帧数 (后续根据文本修改)
# frames_per_image = int(total_frames / num_images)

frames_per_image = []
for result in results:
    frames_per_image.append((int)(total_frames * result / word_sum))

# 创建视频写入对象
output_video = cv2.VideoWriter('output_video_with_fade.mp4', cv2.VideoWriter_fourcc(*'mp4v'), fps,
                               (int(cap.get(3)), int(cap.get(4))))

alpha = 0.9  # 初始透明度
fade_frames = 5  # 渐变帧数
video_width = int(cap.get(3))
video_height = int(cap.get(4))
frame_num = 0
print(num_images)

for i in range(num_images):
    img = cv2.imread(os.path.join(image_folder, sorted_images[i]))
    img_height, img_width = img.shape[:2]
    img_width = int(video_width * 0.7)
    img_height = int(video_height * 0.7)

    # 调整ppt的大小
    img_resized = cv2.resize(img, (img_width, img_height))

    start_frame = int(frame_num)
    end_frame = int(frame_num + frames_per_image[i])

    while frame_num < end_frame:
        ret, frame = cap.read()
        if not ret:
            break

        # 计算插入位置，左边居中
        y_offset = int((frame.shape[0] - img_height) / 2)
        x_offset = 50

        overlay = frame.copy()
        cv2.addWeighted(img_resized, alpha, overlay[y_offset:y_offset + int(img_height), x_offset:x_offset + int(img_width)], 1 - alpha, 0, frame[y_offset:y_offset + int(img_height), x_offset:x_offset + int(img_width)])

        output_video.write(frame)

        frame_num += 1

    # 渐变效果
    for j in range(fade_frames):
        ret, frame = cap.read()
        if not ret:
            break

        overlay = frame.copy()
        cv2.addWeighted(img_resized, 1 - alpha * (j+1) / fade_frames, overlay[y_offset:y_offset + int(img_height), x_offset:x_offset + int(img_width)], alpha * (j+1) / fade_frames, 0, frame[y_offset:y_offset + int(img_height), x_offset:x_offset + int(img_width)])

        output_video.write(frame)

output_video.release()
cap.release()
cv2.destroyAllWindows()

video_clip = VideoFileClip('output_video_with_fade.mp4')
audio_clip = VideoFileClip(video_file).audio
video_clip = video_clip.set_audio(audio_clip)

# 保存最终的视频文件
video_clip.write_videofile(save_name, codec='libx264')

# 关闭视频和音频文件
video_clip.close()
audio_clip.close()
# 关闭视频和音频文件
video_clip.close()
audio_clip.close()
